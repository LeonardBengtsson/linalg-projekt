from enum import Enum
from typing import Optional, Any
import os

import numpy
from numpy import ndarray

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from data.parse_util import NON_COUNTRIES, SYNONYMS, check_synonym
from data.fetch_util import fetch_workbook_from_url, load_matrix, save_matrix



"""
Fetch and parse data
"""


SPREADSHEET_CACHE_DIRECTORY = '.cache/spreadsheet/'
MATRIX_CACHE_DIRECTORY = '.cache/matrix/'

POPULATION_DATA_URL = 'https://population.un.org/wpp/Download/Files/1_Indicator%20(Standard)/EXCEL_FILES/1_General/WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_COMPACT.xlsx'
POPULATION_FILE_PATH = SPREADSHEET_CACHE_DIRECTORY + 'population.xlsx'
POPULATION_SHEET = 'Estimates'

POPULATION_ROW_OFFSET = 18
POPULATION_MIN_COLUMN = 3
POPULATION_MAX_COLUMN = 13
POPULATION_REGION_COLUMN = 3
POPULATION_YEAR_COLUMN = 11
POPULATION_COUNT_COLUMN = 13

IMS_DATA_URL = 'https://www.un.org/development/desa/pd/sites/www.un.org.development.desa.pd/files/undesa_pd_2020_ims_stock_by_sex_destination_and_origin.xlsx'
IMS_FILE_PATH = SPREADSHEET_CACHE_DIRECTORY + 'ims.xlsx'
IMS_SHEET = 'Table 1'

IMS_ROW_OFFSET = 12
IMS_MIN_COLUMN = 2
IMS_MAX_COLUMN = 14
IMS_DESTINATION_COLUMN = 2
IMS_ORIGIN_COLUMN = 6

COUNTRY_MIN_DATA_POINTS = 10 # the minimum amount of data points required for a country to be included in the data set


class Year(Enum):
    Y1990 = 1990
    Y1995 = 1995
    Y2000 = 2000
    Y2005 = 2005
    Y2010 = 2010
    Y2015 = 2015
    Y2020 = 2020

    def migration_column(self):
        match self:
            case Year.Y1990:
                return 8
            case Year.Y1995:
                return 9
            case Year.Y2000:
                return 10
            case Year.Y2005:
                return 11
            case Year.Y2010:
                return 12
            case Year.Y2015:
                return 13
            case Year.Y2020:
                return 14


class DataSet:
    def __init__(
            self,
            year: Year,
            countries: list[str],
            population_counts: ndarray,
            ims_matrix: ndarray,
    ):
        assert len(countries) == len(population_counts) == len(ims_matrix)
        assert numpy.all(numpy.diagonal(ims_matrix) == 0)
        self.year: Year = year
        self.countries: list[str] = countries
        self.population: ndarray = population_counts
        self.ims_matrix: ndarray = ims_matrix


    def get_population(self, country: str) -> int:
        assert country in self.countries
        index: int = self.countries.index(country)
        return self.population[index]

    def get_ims(self, origin: str, destination: str) -> int:
        assert origin in self.countries and destination in self.countries
        origin_index: int = self.countries.index(origin)
        destination_index: int = self.countries.index(destination)
        return self.ims_matrix[destination_index, origin_index]

    def save(self):
        path: str = MATRIX_CACHE_DIRECTORY + str(self.year.value) + '/'
        countries_path: str = path + 'countries.txt'
        population_path: str = path + 'population.txt'
        ims_path: str = path + 'ims.txt'

        os.makedirs(f'{MATRIX_CACHE_DIRECTORY}{str(self.year.value)}', exist_ok=True)
        numpy.savetxt(countries_path, numpy.array(self.countries), fmt='%s')
        numpy.savetxt(population_path, self.population, fmt='%i')
        numpy.savetxt(ims_path, self.ims_matrix, fmt='%i')


def _fetch_data() -> Optional[tuple[Worksheet, Worksheet]]:
    print('Retrieving population data spreadsheet...')
    population_book: Workbook = fetch_workbook_from_url(POPULATION_DATA_URL, POPULATION_FILE_PATH)
    if not population_book:
        print('Failed to retrieve population data spreadsheet')
        return None

    print('Retrieving ims data spreadsheet...')
    ims_book: Workbook = fetch_workbook_from_url(IMS_DATA_URL, IMS_FILE_PATH)
    if not ims_book:
        print('Failed to retrieve ims data spreadsheet')
        return None

    return population_book[POPULATION_SHEET], ims_book[IMS_SHEET]


def _parse(population_sheet: Worksheet, migration_sheet: Worksheet, target_year: Year) -> Optional[DataSet]:
    print('Parsing data...')

    population: dict[str, int] = {}
    for row in population_sheet.iter_rows(POPULATION_ROW_OFFSET, 22000, POPULATION_MIN_COLUMN, POPULATION_MAX_COLUMN):
        year: int = row[POPULATION_YEAR_COLUMN - POPULATION_MIN_COLUMN].value
        if year != target_year.value:
            continue

        country: str = check_synonym(
            row[POPULATION_REGION_COLUMN - POPULATION_MIN_COLUMN]
                .value.replace('*', '').strip())

        if country.lower() in NON_COUNTRIES:
            continue

        population_count: int = row[POPULATION_COUNT_COLUMN - POPULATION_MIN_COLUMN].value
        population[country] = population_count

    migration: dict[str, dict[str, int]] = {}
    for row in migration_sheet.iter_rows(IMS_ROW_OFFSET, migration_sheet.max_row, IMS_MIN_COLUMN, IMS_MAX_COLUMN):
        origin: str = check_synonym(
            row[IMS_ORIGIN_COLUMN - IMS_MIN_COLUMN]
            .value.replace('*', '').strip())

        destination: str = check_synonym(
            row[IMS_DESTINATION_COLUMN - IMS_MIN_COLUMN]
            .value.replace('*', '').strip())

        if origin.lower() in NON_COUNTRIES or destination.lower() in NON_COUNTRIES:
            continue

        migration_count: int = row[target_year.migration_column() - IMS_MIN_COLUMN].value

        if origin not in migration:
            migration[origin] = {}
        migration[origin][destination] = migration_count

    countries: set[str] = set(population.keys()).intersection(set(migration.keys()))
    excluded_countries: set[str] = set()
    for origin_country, destination_set in migration.items():
        if len(destination_set.keys()) < COUNTRY_MIN_DATA_POINTS:
            countries.discard(origin_country)
            excluded_countries.add(origin_country)

    country_list: list[str] = sorted(countries)
    population_data = numpy.zeros(len(country_list), numpy.uint32)
    migration_data = numpy.zeros((len(country_list), len(country_list)), numpy.uint32)

    for i, c in enumerate(country_list):
        population_data[i] = 1000 * population[c]
    for i, origin in enumerate(country_list):
        for j, destination in enumerate(country_list):
            migration_data[j, i] = migration[origin].get(destination, 0)

    return DataSet(target_year, country_list, population_data, migration_data)


def get_data(year: Year) -> DataSet:
    path: str = MATRIX_CACHE_DIRECTORY + str(year.value) + '/'
    countries_path: str = path + 'countries.txt'
    population_path: str = path + 'population.txt'
    ims_path: str = path + 'ims.txt'


    countries: Optional[ndarray] = None
    if os.path.isfile(countries_path):
        countries = numpy.loadtxt(countries_path, dtype='str', delimiter='$')
    population: Optional[ndarray] = load_matrix(population_path)
    ims: Optional[ndarray] = load_matrix(ims_path)

    if population is None or ims is None:
        population_worksheet, ims_worksheet = _fetch_data()
        assert population_worksheet is not None and ims_worksheet is not None
        dataset = _parse(population_worksheet, ims_worksheet, year)
        dataset.save()
        return dataset

    return DataSet(year, countries.tolist(), population, ims)
