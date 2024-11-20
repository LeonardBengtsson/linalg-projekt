import os.path
from enum import Enum
from typing import Optional, Any

import numpy

import urllib.request
from urllib.request import OpenerDirector

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


"""
Fetch and parse data
"""


FILE_DIRECTORY = '../.cache/'

POPULATION_DATA_URL = 'https://population.un.org/wpp/Download/Files/1_Indicator%20(Standard)/EXCEL_FILES/1_General/WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_COMPACT.xlsx'
POPULATION_FILE_PATH = FILE_DIRECTORY + 'population.xlsx'
POPULATION_SHEET = 'Estimates'

POPULATION_ROW_OFFSET = 18
POPULATION_MIN_COLUMN = 3
POPULATION_MAX_COLUMN = 13
POPULATION_REGION_COLUMN = 3
POPULATION_YEAR_COLUMN = 11
POPULATION_COUNT_COLUMN = 13

MIGRATION_DATA_URL = 'https://www.un.org/development/desa/pd/sites/www.un.org.development.desa.pd/files/undesa_pd_2020_ims_stock_by_sex_destination_and_origin.xlsx'
MIGRATION_FILE_PATH = FILE_DIRECTORY + 'migration.xlsx'
MIGRATION_SHEET = 'Table 1'

MIGRATION_ROW_OFFSET = 12
MIGRATION_MIN_COLUMN = 2
MIGRATION_MAX_COLUMN = 14
MIGRATION_DESTINATION_COLUMN = 2
MIGRATION_ORIGIN_COLUMN = 6

NON_COUNTRIES = [
    'world',
    'sub-saharan africa',
    'northern africa and western asia',
    'central and southern asia',
    'eastern and south-eastern asia',
    'latin america and the caribbean',
    'oceania (excluding australia and new zealand)',
    'australia and new zealand',
    'europe and northern america',
    'developed regions',
    'less developed regions',
    'less developed regions, excluding least developed countries',
    'less developed regions, excluding china',
    'least developed countries',
    'land-locked developing countries (lldc)',
    'small island developing states (sids)',
    'high-income countries',
    'middle-income countries',
    'upper-middle-income countries',
    'lower-middle-income countries',
    'low-income countries',
    'africa',
    'eastern africa',
    'middle africa',
    'northern africa',
    'southern africa',
    'western africa',
    'asia',
    'central asia',
    'eastern asia',
    'south-eastern asia',
    'southern asia',
    'western asia',
    'europe',
    'eastern europe',
    'northern europe',
    'southern europe',
    'western europe',
    'caribbean',
    'central america',
    'south america',
    'northern america',
    'oceania',
    'melanesia',
    'micronesia',
    'polynesia',
    'other',
    'australia/new zealand',
    'europe, northern america, australia, and new zealand',
    'sids atlantic',
    'sids atlantic, indian ocean and south china sea (ais)',
    'lldc: europe',
    'sids pacific',
    'lldc: latin america',
    'low-and-middle-income countries',
    'lldc: asia',
    'lldc: africa',
    'no income group available',
    'sids caribbean',
    'more developed regions',
    'high-and-upper-middle-income countries',
    'low-and-lower-middle-income countries'
]

for i in range(len(NON_COUNTRIES)):
    for j in range(i + 1, len(NON_COUNTRIES)):
        if NON_COUNTRIES[i] == NON_COUNTRIES[j]:
            print(NON_COUNTRIES[i])


SYNONYMS: list[tuple[str, str]] = [
    ('Turkey', 'türkiye'),
    ('Kosovo', 'kosovo (under unsc res. 1244)'),
    ('Taiwan', 'china, taiwan province of china'),
    ('Bolivia', 'bolivia (plurinational state of)'),
    ('Brunei', 'brunei darussalam'),
    ('Hong Kong', 'China, Hong Kong SAR'),
    ('Macao', 'China, Macao SAR'),
    ('North Korea', 'Dem. People\'s Republic of Korea'),
    ('Vatican City', 'Holy See'),
    ('Iran', 'Iran (Islamic Republic of)'),
    ('Laos', 'Lao People\'s Democratic Republic'),
    ('Federated States of Micronesia', 'Micronesia (Fed. States of)'),
    ('Russia', 'Russian Federation'),
    ('Syria', 'Syrian Arab Republic'),
    ('Tanzania', 'United Republic of Tanzania'),
    ('Venezuela', 'Venezuela (Bolivarian Republic of)'),
    ('Vietnam', 'Viet Nam')
]


def _check_synonym(country: str) -> str:
    for a, b in SYNONYMS:
        if country.lower() == b.lower():
            return a
    return country


class Year(Enum):
    Y1990 = 1990
    Y1995 = 1995
    Y2000 = 2000
    Y2005 = 2005
    Y2010 = 2010
    Y2015 = 2015
    Y2020 = 2020

    def migration_column(self):
        match self:
            case Year.Y1990:
                return 8
            case Year.Y1995:
                return 9
            case Year.Y2000:
                return 10
            case Year.Y2005:
                return 11
            case Year.Y2010:
                return 12
            case Year.Y2015:
                return 13
            case Year.Y2020:
                return 14


class DataSet:
    def __init__(self, countries: list[str], population_counts: numpy.ndarray[Any, Any], migration_matrix: numpy.ndarray[Any, Any]):
        assert len(countries) == len(population_counts) == len(migration_matrix)
        self.countries: list[str] = countries
        self._population_counts: numpy.ndarray[Any, Any] = population_counts
        self._migration_matrix: numpy.ndarray[Any, Any] = migration_matrix

    def get_population(self, country: str) -> int:
        assert country in self.countries
        index: int = self.countries.index(country)
        return self._population_counts[index]

    def get_migration(self, origin: str, destination: str) -> int:
        assert origin in self.countries
        origin_index: int = self.countries.index(origin)
        assert destination in self.countries
        destination_index: int = self.countries.index(destination)
        return self._migration_matrix[origin_index][destination_index]


def _fetch_from_url(url: str, path: str) -> Optional[Workbook]:
    if not os.path.isfile(path):
        directory: str = os.path.dirname(path)
        if not os.path.isdir(directory):
            os.mkdir(directory)

        opener: OpenerDirector = urllib.request.build_opener()
        opener.addheaders = [('User-agent', '')]
        urllib.request.install_opener(opener)
        try:
            urllib.request.urlretrieve(url, path)
        except:
            return None
    if not os.path.isfile(path):
        return None
    return openpyxl.open(path, read_only=True)


def _fetch_data() -> Optional[tuple[Worksheet, Worksheet]]:
    print('Retrieving population data spreadsheet...')
    population_book: Workbook = _fetch_from_url(POPULATION_DATA_URL, POPULATION_FILE_PATH)
    if not population_book:
        print('Failed to retrieve population data spreadsheet')
        return None

    print('Retrieving migration data spreadsheet...')
    migration_book: Workbook = _fetch_from_url(MIGRATION_DATA_URL, MIGRATION_FILE_PATH)
    if not migration_book:
        print('Failed to retrieve migration data spreadsheet')
        return None

    return population_book[POPULATION_SHEET], migration_book[MIGRATION_SHEET]


def parse(population_sheet: Worksheet, migration_sheet: Worksheet, target_year: Year) -> Optional[DataSet]:
    print('Parsing data...')

    population: dict[str, int] = {}
    for row in population_sheet.iter_rows(POPULATION_ROW_OFFSET, 22000, POPULATION_MIN_COLUMN, POPULATION_MAX_COLUMN):
        year: int = row[POPULATION_YEAR_COLUMN - POPULATION_MIN_COLUMN].value
        if year != target_year.value:
            continue

        country: str = _check_synonym(
            row[POPULATION_REGION_COLUMN - POPULATION_MIN_COLUMN]
                .value.replace('*', '').strip())

        if country.lower() in NON_COUNTRIES:
            continue

        population_count: int = row[POPULATION_COUNT_COLUMN - POPULATION_MIN_COLUMN].value
        population[country] = population_count

    migration: dict[str, dict[str, int]] = {}
    for row in migration_sheet.iter_rows(MIGRATION_ROW_OFFSET, migration_sheet.max_row, MIGRATION_MIN_COLUMN, MIGRATION_MAX_COLUMN):
        origin: str = _check_synonym(
            row[MIGRATION_ORIGIN_COLUMN - MIGRATION_MIN_COLUMN]
            .value.replace('*', '').strip())

        destination: str = _check_synonym(
            row[MIGRATION_DESTINATION_COLUMN - MIGRATION_MIN_COLUMN]
            .value.replace('*', '').strip())

        if origin.lower() in NON_COUNTRIES or destination.lower() in NON_COUNTRIES:
            continue

        migration_count: int = row[target_year.migration_column() - MIGRATION_MIN_COLUMN].value

        if origin not in migration:
            migration[origin] = {}
        migration[origin][destination] = migration_count

    countries: set[str] = set(population.keys()).union(set(migration.keys()))
    _missed_countries: set[str] = set()
    for origin_country in migration.values():
        _missed_countries.update(set(origin_country.keys()).difference(countries))
        print(origin_country.keys())
        countries = countries.intersection(set(origin_country.keys()))
    print(sorted(countries))
    print(sorted(_missed_countries))

    country_list: list[str] = sorted(countries)
    population_data = numpy.zeros(len(country_list), numpy.uint32)
    migration_data = numpy.zeros((len(country_list), len(country_list)), numpy.uint32)

    country_iter: enumerate[str] = enumerate(country_list)
    for i, c in country_iter:
        population_data[i] = population[c]
    for i, origin in country_iter:
        for j, destination in country_iter:
            migration_data[i][j] = migration[origin][destination]

    return DataSet(country_list, population_data, migration_data)


if __name__ == '__main__':
    population_data, migration_data = _fetch_data()
    assert population_data is not None and migration_data is not None

    dataset_2010: DataSet = parse(population_data, migration_data, Year.Y2010)
    assert dataset_2010 is not None

    dataset_2020: DataSet = parse(population_data, migration_data, Year.Y2020)
    assert dataset_2020 is not None

    print('Norway -> Sweden', dataset_2020.get_migration('Norway', 'Sweden') - dataset_2010.get_migration('Norway', 'Sweden'))
